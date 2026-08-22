"""Deterministic, offline-first extraction for annuity product materials.

The module deliberately does not perform OCR or make network calls. Native PDF
text and tables are preferred, ``pdftotext`` is a local fallback, and pages that
still have no machine-readable content are placed in a targeted OCR queue.
Customer-specific data is rejected and redacted before extraction or review
artifacts are written.
"""

from __future__ import annotations

import csv
import hashlib
import importlib.metadata
import io
import json
import math
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import unicodedata
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any


SCHEMA_VERSION = "1.0.0"
EXTRACTOR_VERSION = "1.0.0"
_STDLIB_VERSION = (
    f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
)

_DEFAULT_KEYWORDS: dict[str, tuple[str, ...]] = {
    "annuity_cash_flow": (
        "年金",
        "养老金",
        "养老年金",
        "领取金额",
        "开始领取",
        "annuity",
        "pension",
        "income payment",
        "benefit payment",
    ),
    "premium_structure": (
        "保费",
        "保险费",
        "缴费期",
        "交费期",
        "趸交",
        "premium",
        "single premium",
        "payment term",
    ),
    "cash_value_liquidity": (
        "现金价值",
        "退保",
        "保单贷款",
        "减保",
        "cash value",
        "surrender",
        "policy loan",
        "liquidity",
    ),
    "death_benefit": (
        "身故保险金",
        "死亡保险金",
        "身故给付",
        "death benefit",
        "death payment",
    ),
    "longevity_guarantee": (
        "终身领取",
        "保证领取",
        "生存保险金",
        "领取期间",
        "lifetime income",
        "life annuity",
        "guaranteed period",
        "survival benefit",
    ),
    "inflation_escalation": (
        "通胀",
        "递增领取",
        "年增",
        "领取递增",
        "inflation",
        "escalating annuity",
        "increasing annuity",
        "cost of living",
    ),
    "guarantee_status": (
        "保证利益",
        "非保证利益",
        "红利",
        "分红",
        "万能账户",
        "guaranteed benefit",
        "non-guaranteed",
        "dividend",
        "bonus",
    ),
    "product_eligibility": (
        "投保年龄",
        "承保年龄",
        "保险期间",
        "issue age",
        "entry age",
        "policy term",
    ),
}

_SEMANTIC_AMBIGUITY_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    (
        "greater_or_lesser_branch",
        re.compile(
            r"(?:取|以)(?:二者|三者|以上|下列|前述)?.{0,24}(?:较大|较小|最大|最小)者"
            r"|whichever\s+is\s+(?:greater|higher|lesser|lower)",
            re.IGNORECASE,
        ),
    ),
    (
        "cross_reference_requires_interpretation",
        re.compile(
            r"另有约定|按(?:本合同|条款)约定(?:执行|确定)|除非.{0,30}约定"
            r"|subject\s+to\s+(?:the\s+)?(?:policy|contract|terms?)"
            r"|except\s+as\s+(?:otherwise\s+)?provided",
            re.IGNORECASE,
        ),
    ),
)

_PROHIBITED_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    (
        "customer_field",
        re.compile(
            r"(?:(?:客户|投保人|被保险人|受益人)\s*"
            r"(?:姓名|身份证(?:号|号码)?|证件(?:号|号码)?|手机号(?:码)?|联系电话|邮箱|"
            r"住址|地址|出生日期|年龄(?!\s*(?:范围|区间|限制))|收入|资产|净资产|职业|健康状况|家庭结构|"
            r"退休支出|风险(?:承受能力|偏好)|传承(?:目标|需求)|期望领取(?:金额)?|"
            r"希望领取(?:金额)?)"
            r"|(?:customer|client|applicant|insured|beneficiary)\s*"
            r"(?:name|id(?:entification)?(?:\s+number)?|passport(?:\s+number)?|phone|"
            r"mobile|e-?mail|address|date\s+of\s+birth|age(?!\s*(?:range|limits?))|income|assets?|net\s+worth|"
            r"occupation|health|family|retirement\s+spend(?:ing)?|risk\s+tolerance|"
            r"legacy\s+goal))\s*[:：=]\s*[^\r\n,，;；|]{0,96}",
            re.IGNORECASE,
        ),
    ),
    (
        "email_address",
        re.compile(
            r"(?<![\w.+-])[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}(?![\w.-])",
            re.IGNORECASE,
        ),
    ),
    (
        "mainland_china_phone",
        re.compile(r"(?<!\d)(?:\+?86[-\s]?)?1[3-9]\d{9}(?!\d)"),
    ),
    (
        "mainland_china_id",
        re.compile(r"(?<![0-9A-Z])\d{17}[0-9X](?![0-9A-Z])", re.IGNORECASE),
    ),
    (
        "payment_card_or_account",
        re.compile(r"(?<![0-9A-Za-z])(?:\d[ -]?){15,18}\d(?![0-9A-Za-z])"),
    ),
)

_SUPPORTED_SUFFIXES = {".csv", ".tsv", ".xlsx", ".json", ".txt", ".md", ".pdf"}


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_text(value: str) -> str:
    return _sha256_bytes(value.encode("utf-8"))


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def _write_stable_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rendered = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        indent=2,
        allow_nan=False,
    )
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(rendered + "\n")


def _module_version(distribution_name: str, fallback: str = "unknown") -> str:
    try:
        return importlib.metadata.version(distribution_name)
    except importlib.metadata.PackageNotFoundError:
        return fallback


def fingerprint_file(path: os.PathLike[str] | str) -> dict[str, Any]:
    """Return a content-based fingerprint without changing the input file.

    Modification times are intentionally excluded so identical bytes at the
    same path produce the same result across repeated inspections.
    """

    resolved = Path(os.fspath(path)).expanduser().resolve()
    if not resolved.exists():
        raise FileNotFoundError(str(resolved))
    if not resolved.is_file():
        raise ValueError(f"Expected a file, got: {resolved}")

    digest = hashlib.sha256()
    with resolved.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)

    media_type, _ = mimetypes.guess_type(resolved.name)
    return {
        "extension": resolved.suffix.lower(),
        "media_type": media_type or "application/octet-stream",
        "name": resolved.name,
        "path": str(resolved),
        "sha256": digest.hexdigest(),
        "size_bytes": resolved.stat().st_size,
    }


def detect_prohibited_customer_data(text: str) -> list[dict[str, Any]]:
    """Locate probable customer-specific data without returning matched values.

    Product-level fields such as ``投保年龄范围`` and ``issue age`` are not
    prohibited. Explicit customer/applicant fields and common direct identifiers
    are. Each finding retains only its span and a one-way hash for auditability.
    """

    if not isinstance(text, str):
        raise TypeError("text must be a string")

    findings: list[dict[str, Any]] = []
    seen: set[tuple[str, int, int]] = set()
    for finding_type, pattern in _PROHIBITED_PATTERNS:
        for match in pattern.finditer(text):
            key = (finding_type, match.start(), match.end())
            if key in seen:
                continue
            seen.add(key)
            findings.append(
                {
                    "end": match.end(),
                    "match_sha256": _sha256_text(match.group(0)),
                    "start": match.start(),
                    "type": finding_type,
                }
            )
    stripped = unicodedata.normalize("NFKC", text).strip()
    sensitive_label = re.fullmatch(
        r"(?:(?:客户|投保人|被保险人|受益人)\s*(?:姓名|身份证(?:号|号码)?|证件(?:号|号码)?|"
        r"手机号(?:码)?|联系电话|邮箱|住址|地址|出生日期|收入|资产|净资产|健康状况|"
        r"家庭结构|风险(?:承受能力|偏好)|传承(?:目标|需求)|期望领取(?:金额)?|希望领取(?:金额)?)"
        r"|(?:customer|client|applicant|insured|beneficiary)\s*(?:name|id(?:entification)?(?:\s+number)?|"
        r"passport(?:\s+number)?|phone|mobile|e-?mail|address|date\s+of\s+birth|income|assets?|"
        r"net\s+worth|health|family|risk\s+tolerance|legacy\s+goal))",
        stripped,
        re.IGNORECASE,
    )
    if sensitive_label:
        key = ("customer_field", sensitive_label.start(), sensitive_label.end())
        if key not in seen:
            findings.append(
                {
                    "end": sensitive_label.end(),
                    "match_sha256": _sha256_text(sensitive_label.group(0)),
                    "start": sensitive_label.start(),
                    "type": "customer_field",
                }
            )
    if re.fullmatch(
        r"客户\s*年龄|投保人\s*职业|customer\s*age|applicant\s*occupation",
        stripped,
        re.IGNORECASE,
    ):
        key = ("customer_field", 0, len(stripped))
        if key not in seen:
            findings.append(
                {
                    "end": len(stripped),
                    "match_sha256": _sha256_text(stripped),
                    "start": 0,
                    "type": "customer_field",
                }
            )
    findings.sort(key=lambda item: (item["start"], item["end"], item["type"]))
    return findings


def route_confidence(
    score: float,
    has_conflict: bool,
    semantic_ambiguity: bool,
    image_only: bool,
) -> str:
    """Route evidence without allowing low confidence alone to invoke an LLM.

    Numeric conflicts always require deterministic/manual verification. Image-
    only evidence requires page-targeted OCR. The LLM route is reserved solely
    for text that has already been identified as semantically ambiguous.
    """

    try:
        numeric_score = float(score)
    except (TypeError, ValueError) as exc:
        raise ValueError("score must be a finite number between 0 and 1") from exc
    if not math.isfinite(numeric_score) or not 0.0 <= numeric_score <= 1.0:
        raise ValueError("score must be a finite number between 0 and 1")

    if has_conflict:
        return "deterministic_manual_verification"
    if image_only:
        return "targeted_ocr"
    if semantic_ambiguity and numeric_score >= 0.90:
        return "llm_semantic_resolution"
    if numeric_score >= 0.90:
        return "direct_accept"
    if numeric_score >= 0.70:
        return "deterministic_second_pass"
    return "manual_verification"


def _decode_text_file(path: Path) -> tuple[str, str]:
    data = path.read_bytes()
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        return data.decode("utf-16"), "utf-16"
    for encoding in ("utf-8-sig", "gb18030", "latin-1"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    # latin-1 is total, so this is defensive rather than reachable.
    return data.decode("utf-8", errors="replace"), "utf-8-replacement"


def _extractor(name: str, version: str, **extra: Any) -> dict[str, Any]:
    value: dict[str, Any] = {"name": name, "version": version}
    value.update(extra)
    return value


def _raw_record(
    raw_text: str,
    *,
    record_type: str,
    location: dict[str, Any],
    extractor: dict[str, Any],
    confidence: float,
) -> dict[str, Any]:
    return {
        "confidence": confidence,
        "extractor": extractor,
        "location": location,
        "raw_text": raw_text,
        "record_type": record_type,
    }


def _extract_delimited(
    path: Path, delimiter: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    text, encoding = _decode_text_file(path)
    extractor = _extractor(
        "python.csv",
        _STDLIB_VERSION,
        delimiter=delimiter,
        encoding=encoding,
    )
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    try:
        reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
        for row_index, row in enumerate(reader, start=1):
            for column_index, value in enumerate(row, start=1):
                if not value.strip():
                    continue
                records.append(
                    _raw_record(
                        value,
                        record_type="table_cell",
                        location={
                            "cell": _excel_column(column_index) + str(row_index),
                            "column": column_index,
                            "row": row_index,
                            "sheet": path.stem,
                        },
                        extractor=extractor,
                        confidence=0.995,
                    )
                )
    except csv.Error:
        issues.append(
            {
                "confidence": 0.0,
                "kind": "delimited_parse_error",
                "reason": "Delimited text could not be parsed deterministically.",
            }
        )
    return records, issues


def _extract_text(
    path: Path, record_type: str = "text_line"
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    text, encoding = _decode_text_file(path)
    extractor = _extractor("python.text", _STDLIB_VERSION, encoding=encoding)
    records = [
        _raw_record(
            line.rstrip("\r"),
            record_type=record_type,
            location={"line": line_number},
            extractor=extractor,
            confidence=0.995,
        )
        for line_number, line in enumerate(text.splitlines(), start=1)
        if line.strip()
    ]
    issues: list[dict[str, Any]] = []
    if not records:
        issues.append(
            {
                "confidence": 0.0,
                "kind": "empty_document",
                "reason": "No non-empty text was found.",
            }
        )
    return records, issues


def _json_pointer_token(value: str) -> str:
    return value.replace("~", "~0").replace("/", "~1")


def _extract_json(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    text, encoding = _decode_text_file(path)
    extractor = _extractor("python.json", _STDLIB_VERSION, encoding=encoding)
    try:
        payload = json.loads(text)
    except (json.JSONDecodeError, UnicodeError):
        records, _ = _extract_text(path, record_type="json_fallback_line")
        for record in records:
            record["confidence"] = 0.50
            record["extractor"] = _extractor(
                "python.json_text_fallback",
                _STDLIB_VERSION,
                encoding=encoding,
            )
        return records, [
            {
                "confidence": 0.0,
                "kind": "json_parse_error",
                "reason": "JSON syntax is invalid; text fallback is non-authoritative.",
            }
        ]

    records: list[dict[str, Any]] = []

    def walk(value: Any, pointer: str, label: str) -> None:
        if isinstance(value, Mapping):
            for key in sorted(value, key=lambda item: str(item)):
                key_text = str(key)
                walk(
                    value[key],
                    pointer + "/" + _json_pointer_token(key_text),
                    key_text,
                )
            return
        if isinstance(value, list):
            for index, item in enumerate(value):
                walk(item, pointer + "/" + str(index), label)
            return

        scalar = json.dumps(
            value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
        raw_text = f"{label}: {scalar}" if label else scalar
        records.append(
            _raw_record(
                raw_text,
                record_type="json_scalar",
                location={"json_pointer": pointer or "/"},
                extractor=extractor,
                confidence=0.995,
            )
        )

    walk(payload, "", "")
    return records, []


def _excel_column(index: int) -> str:
    result = ""
    value = index
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _stable_cell_text(value: Any) -> str:
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _extract_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError:
        return [], [
            {
                "confidence": 0.0,
                "kind": "optional_dependency_missing",
                "reason": "openpyxl is required to inspect XLSX files locally.",
            }
        ]

    extractor = _extractor(
        "openpyxl", getattr(openpyxl, "__version__", "unknown"), data_only=False
    )
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    except Exception:  # openpyxl exposes several format-specific exception classes.
        return [], [
            {
                "confidence": 0.0,
                "kind": "xlsx_parse_error",
                "reason": "The XLSX package could not be parsed deterministically.",
            }
        ]

    try:
        for sheet_name in sorted(workbook.sheetnames):
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None or str(cell.value).strip() == "":
                        continue
                    records.append(
                        _raw_record(
                            _stable_cell_text(cell.value),
                            record_type="workbook_cell",
                            location={
                                "cell": cell.coordinate,
                                "column": cell.column,
                                "row": cell.row,
                                "sheet": sheet_name,
                            },
                            extractor=extractor,
                            confidence=0.995,
                        )
                    )
    finally:
        workbook.close()

    if not records:
        issues.append(
            {
                "confidence": 0.0,
                "kind": "empty_workbook",
                "reason": "No populated XLSX cells were found.",
            }
        )
    return records, issues


def _rounded_bbox(values: Sequence[Any]) -> list[float]:
    return [round(float(value), 3) for value in values]


def _group_pdf_words(
    words: list[dict[str, Any]], tolerance: float = 3.0
) -> list[tuple[str, list[float]]]:
    ordered = sorted(
        words,
        key=lambda item: (float(item.get("top", 0.0)), float(item.get("x0", 0.0))),
    )
    groups: list[list[dict[str, Any]]] = []
    for word in ordered:
        if (
            not groups
            or abs(float(word.get("top", 0.0)) - float(groups[-1][0].get("top", 0.0)))
            > tolerance
        ):
            groups.append([word])
        else:
            groups[-1].append(word)

    lines: list[tuple[str, list[float]]] = []
    for group in groups:
        group.sort(key=lambda item: float(item.get("x0", 0.0)))
        text = " ".join(str(item.get("text", "")) for item in group).strip()
        if not text:
            continue
        bbox = [
            min(float(item.get("x0", 0.0)) for item in group),
            min(float(item.get("top", 0.0)) for item in group),
            max(float(item.get("x1", 0.0)) for item in group),
            max(float(item.get("bottom", 0.0)) for item in group),
        ]
        lines.append((text, _rounded_bbox(bbox)))
    return lines


def _extract_pdfplumber(
    path: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int, bool]:
    try:
        import pdfplumber  # type: ignore[import-not-found]
    except ImportError:
        return [], [], 0, False

    version = getattr(pdfplumber, "__version__", _module_version("pdfplumber"))
    extractor = _extractor("pdfplumber", version, mode="native_words_and_tables")
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    page_count = 0
    try:
        with pdfplumber.open(path) as document:
            page_count = len(document.pages)
            for page_number, page in enumerate(document.pages, start=1):
                page_record_count = 0
                try:
                    tables = page.find_tables()
                    for table_index, table in enumerate(tables, start=1):
                        table_bbox = _rounded_bbox(table.bbox)
                        for row_index, row in enumerate(table.extract(), start=1):
                            for column_index, value in enumerate(row, start=1):
                                if value is None or not str(value).strip():
                                    continue
                                records.append(
                                    _raw_record(
                                        str(value),
                                        record_type="pdf_table_cell",
                                        location={
                                            "bbox": table_bbox,
                                            "column": column_index,
                                            "page": page_number,
                                            "row": row_index,
                                            "table": table_index,
                                        },
                                        extractor=extractor,
                                        confidence=0.95,
                                    )
                                )
                                page_record_count += 1
                except Exception:
                    issues.append(
                        {
                            "confidence": 0.60,
                            "kind": "pdf_table_extraction_error",
                            "location": {"page": page_number},
                            "reason": "Native table extraction failed for this page.",
                        }
                    )

                try:
                    words = page.extract_words(
                        x_tolerance=2,
                        y_tolerance=3,
                        keep_blank_chars=False,
                        use_text_flow=False,
                    )
                except Exception:
                    words = []
                    issues.append(
                        {
                            "confidence": 0.0,
                            "kind": "pdf_word_extraction_error",
                            "location": {"page": page_number},
                            "reason": "Native word extraction failed for this page.",
                        }
                    )

                for line_number, (text, bbox) in enumerate(
                    _group_pdf_words(words), start=1
                ):
                    records.append(
                        _raw_record(
                            text,
                            record_type="pdf_text_line",
                            location={
                                "bbox": bbox,
                                "line": line_number,
                                "page": page_number,
                            },
                            extractor=extractor,
                            confidence=0.98,
                        )
                    )
                    page_record_count += 1

                if page_record_count == 0:
                    issues.append(
                        {
                            "confidence": 0.0,
                            "image_only": True,
                            "kind": "image_only_page",
                            "location": {"page": page_number},
                            "reason": "No native words or table cells were found; OCR must target only this page.",
                        }
                    )
    except Exception:
        return (
            [],
            [
                {
                    "confidence": 0.0,
                    "kind": "pdf_native_parse_error",
                    "reason": "The PDF could not be opened by the native parser.",
                }
            ],
            page_count,
            True,
        )
    return records, issues, page_count, True


def _pdftotext_version(executable: str) -> str:
    try:
        completed = subprocess.run(
            [executable, "-v"],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=10,
        )
    except (OSError, subprocess.SubprocessError):
        return "unknown"
    message = (
        (completed.stderr or completed.stdout).decode("utf-8", errors="replace").strip()
    )
    return message.splitlines()[0] if message else "unknown"


def _extract_pdftotext(
    path: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int, bool]:
    executable = shutil.which("pdftotext")
    if not executable:
        return [], [], 0, False
    extractor = _extractor(
        "pdftotext", _pdftotext_version(executable), mode="layout_fallback"
    )
    try:
        completed = subprocess.run(
            [executable, "-layout", "-enc", "UTF-8", str(path), "-"],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120,
        )
    except (OSError, subprocess.SubprocessError):
        return (
            [],
            [
                {
                    "confidence": 0.0,
                    "kind": "pdftotext_execution_error",
                    "reason": "The local pdftotext fallback did not complete.",
                }
            ],
            0,
            True,
        )
    if completed.returncode != 0:
        return (
            [],
            [
                {
                    "confidence": 0.0,
                    "kind": "pdftotext_parse_error",
                    "reason": "The local pdftotext fallback could not parse the PDF.",
                }
            ],
            0,
            True,
        )

    text = completed.stdout.decode("utf-8", errors="replace")
    pages = text.split("\f")
    if pages and not pages[-1].strip():
        pages.pop()
    records: list[dict[str, Any]] = []
    for page_number, page_text in enumerate(pages, start=1):
        for line_number, line in enumerate(page_text.splitlines(), start=1):
            if not line.strip():
                continue
            records.append(
                _raw_record(
                    line.rstrip(),
                    record_type="pdf_text_line_fallback",
                    location={"line": line_number, "page": page_number},
                    extractor=extractor,
                    confidence=0.89,
                )
            )
    return records, [], len(pages), True


def _extract_pdf(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    native_records, native_issues, native_page_count, native_available = (
        _extract_pdfplumber(path)
    )
    missing_pages = {
        int(issue["location"]["page"])
        for issue in native_issues
        if issue.get("image_only") and issue.get("location", {}).get("page") is not None
    }

    fallback_records: list[dict[str, Any]] = []
    fallback_issues: list[dict[str, Any]] = []
    fallback_page_count = 0
    fallback_available = False
    if not native_records or missing_pages or not native_available:
        fallback_records, fallback_issues, fallback_page_count, fallback_available = (
            _extract_pdftotext(path)
        )

    if not native_records:
        records = fallback_records
    else:
        recovered = [
            record
            for record in fallback_records
            if int(record.get("location", {}).get("page", -1)) in missing_pages
        ]
        records = native_records + recovered

    recovered_pages = {
        int(record["location"]["page"])
        for record in fallback_records
        if record.get("location", {}).get("page") is not None
    }
    issues = [
        issue
        for issue in native_issues
        if not (
            issue.get("image_only")
            and issue.get("location", {}).get("page") in recovered_pages
        )
    ]
    issues.extend(fallback_issues)

    if not records and not any(issue.get("image_only") for issue in issues):
        known_page_count = max(native_page_count, fallback_page_count)
        if known_page_count:
            for page_number in range(1, known_page_count + 1):
                issues.append(
                    {
                        "confidence": 0.0,
                        "image_only": True,
                        "kind": "image_only_page",
                        "location": {"page": page_number},
                        "reason": "No native or pdftotext content was found; OCR must target only this page.",
                    }
                )
        elif not native_available and not fallback_available:
            issues.append(
                {
                    "confidence": 0.0,
                    "kind": "optional_dependency_missing",
                    "reason": "Install pdfplumber or the local pdftotext utility to inspect PDFs.",
                }
            )
        else:
            issues.append(
                {
                    "confidence": 0.0,
                    "kind": "pdf_content_unresolved",
                    "reason": "No machine-readable PDF content was recovered; determine page scope before OCR.",
                }
            )
    return records, issues


def _extract_path(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _extract_delimited(path, ",")
    if suffix == ".tsv":
        return _extract_delimited(path, "\t")
    if suffix == ".xlsx":
        return _extract_xlsx(path)
    if suffix == ".json":
        return _extract_json(path)
    if suffix in {".txt", ".md"}:
        return _extract_text(path)
    if suffix == ".pdf":
        return _extract_pdf(path)
    return [], [
        {
            "confidence": 0.0,
            "kind": "unsupported_format",
            "reason": f"Unsupported extension: {suffix or '[none]'}",
        }
    ]


def _normalize_keyword(value: str) -> str:
    return unicodedata.normalize("NFKC", value).casefold().strip()


def _keyword_map(keywords: Any) -> dict[str, tuple[str, ...]]:
    result: dict[str, set[str]] = {
        category: {_normalize_keyword(term) for term in terms if str(term).strip()}
        for category, terms in _DEFAULT_KEYWORDS.items()
    }
    if keywords is None:
        pass
    elif isinstance(keywords, Mapping):
        for category, terms in keywords.items():
            if isinstance(terms, str):
                terms = [terms]
            result.setdefault(str(category), set()).update(
                _normalize_keyword(str(term)) for term in terms if str(term).strip()
            )
    elif isinstance(keywords, Sequence) and not isinstance(
        keywords, (str, bytes, bytearray)
    ):
        result.setdefault("custom", set()).update(
            _normalize_keyword(str(term)) for term in keywords if str(term).strip()
        )
    else:
        raise TypeError("keywords must be a mapping, a sequence of strings, or None")
    return {
        category: tuple(sorted(terms))
        for category, terms in sorted(result.items())
        if terms
    }


def _classify_text(
    text: str, keyword_map: Mapping[str, Sequence[str]]
) -> list[dict[str, Any]]:
    normalized = _normalize_keyword(text)
    classifications: list[dict[str, Any]] = []
    for category in sorted(keyword_map):
        matches = sorted(
            {term for term in keyword_map[category] if term and term in normalized}
        )
        if matches:
            classifications.append({"category": category, "matched_keywords": matches})
    return classifications


def _semantic_ambiguity(text: str) -> str | None:
    normalized = unicodedata.normalize("NFKC", text)
    for reason, pattern in _SEMANTIC_AMBIGUITY_PATTERNS:
        if pattern.search(normalized):
            return reason
    return None


def _safe_review_location(location: Mapping[str, Any] | None) -> dict[str, Any]:
    if not location:
        return {}
    safe_keys = {
        "bbox",
        "cell",
        "column",
        "json_pointer",
        "line",
        "page",
        "row",
        "sheet",
        "table",
    }
    return {key: location[key] for key in sorted(location) if key in safe_keys}


def _review_id(value: Mapping[str, Any]) -> str:
    return _sha256_text(_canonical_json(value))[:24]


def _review_item(
    *,
    source_sha256: str,
    route: str,
    reason: str,
    location: Mapping[str, Any] | None = None,
    record_id: str | None = None,
    content_sha256: str | None = None,
    snippet: str | None = None,
    detection_types: Sequence[str] | None = None,
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "location": _safe_review_location(location),
        "reason": reason,
        "route": route,
        "source_sha256": source_sha256,
    }
    if record_id:
        item["record_id"] = record_id
    if content_sha256:
        item["content_sha256"] = content_sha256
    if snippet is not None:
        # This field is only valid for the semantic LLM route.
        if route != "llm_semantic_resolution":
            raise ValueError("Only semantic LLM review items may contain snippets")
        item["snippet"] = snippet
    if detection_types:
        item["detection_types"] = sorted(set(detection_types))
    item["id"] = _review_id(item)
    return item


def _enrich_record(
    raw_record: Mapping[str, Any],
    fingerprint: Mapping[str, Any],
    keyword_map: Mapping[str, Sequence[str]],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    raw_text = str(raw_record.get("raw_text", ""))
    raw_text_sha256 = _sha256_text(raw_text)
    location = dict(raw_record.get("location", {}))
    pii_findings = detect_prohibited_customer_data(raw_text)
    classifications = _classify_text(raw_text, keyword_map)
    identity = {
        "location": location,
        "raw_text_sha256": raw_text_sha256,
        "record_type": raw_record["record_type"],
        "source_sha256": fingerprint["sha256"],
    }
    record_id = _sha256_text(_canonical_json(identity))[:24]
    record = {
        "confidence": float(raw_record["confidence"]),
        "extractor": dict(raw_record["extractor"]),
        "id": record_id,
        "location": location,
        "prohibited_customer_data": pii_findings,
        "raw_text": "[REDACTED:PROHIBITED_CUSTOMER_DATA]" if pii_findings else raw_text,
        "raw_text_redacted": bool(pii_findings),
        "raw_text_sha256": raw_text_sha256,
        "record_type": raw_record["record_type"],
        "section_classifications": classifications,
        "source_name": fingerprint["name"],
        "source_path": fingerprint["path"],
        "source_sha256": fingerprint["sha256"],
    }

    review_items: list[dict[str, Any]] = []
    if pii_findings:
        review_items.append(
            _review_item(
                source_sha256=str(fingerprint["sha256"]),
                route="reject_prohibited_customer_data",
                reason="Customer-specific data is outside this product-only skill.",
                location=location,
                record_id=record_id,
                content_sha256=raw_text_sha256,
                detection_types=[finding["type"] for finding in pii_findings],
            )
        )
        return record, review_items

    semantic_reason = _semantic_ambiguity(raw_text)
    route = route_confidence(
        float(raw_record["confidence"]),
        has_conflict=False,
        semantic_ambiguity=semantic_reason is not None,
        image_only=False,
    )
    if route == "llm_semantic_resolution":
        review_items.append(
            _review_item(
                source_sha256=str(fingerprint["sha256"]),
                route=route,
                reason=str(semantic_reason),
                location=location,
                record_id=record_id,
                content_sha256=raw_text_sha256,
                snippet=raw_text[:1200],
            )
        )
    elif route != "direct_accept":
        review_items.append(
            _review_item(
                source_sha256=str(fingerprint["sha256"]),
                route=route,
                reason="Extraction confidence requires non-LLM verification.",
                location=location,
                record_id=record_id,
                content_sha256=raw_text_sha256,
            )
        )
    return record, review_items


def _issue_review_item(issue: Mapping[str, Any], source_sha256: str) -> dict[str, Any]:
    kind = str(issue.get("kind", "extraction_issue"))
    is_conflict = kind == "numeric_conflict" or bool(issue.get("has_conflict"))
    is_image_only = bool(issue.get("image_only"))
    route = route_confidence(
        float(issue.get("confidence", 0.0)),
        has_conflict=is_conflict,
        semantic_ambiguity=False,
        image_only=is_image_only,
    )
    return _review_item(
        source_sha256=source_sha256,
        route=route,
        reason=str(issue.get("reason", kind)),
        location=issue.get("location", {}),
    )


def _is_within(path: Path, directory: Path) -> bool:
    try:
        path.relative_to(directory)
        return True
    except ValueError:
        return False


def _collect_input_files(
    paths: Sequence[os.PathLike[str] | str] | os.PathLike[str] | str,
    output_dir: Path,
) -> list[Path]:
    if isinstance(paths, (str, os.PathLike)):
        supplied = [paths]
    else:
        supplied = list(paths)
    collected: dict[str, Path] = {}
    for supplied_path in supplied:
        path = Path(os.fspath(supplied_path)).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(str(path))
        candidates = (
            [path]
            if path.is_file()
            else sorted(item for item in path.rglob("*") if item.is_file())
        )
        for candidate in candidates:
            resolved = candidate.resolve()
            if _is_within(resolved, output_dir):
                continue
            collected[str(resolved)] = resolved
    return [collected[key] for key in sorted(collected)]


def inspect_inputs(
    paths: Sequence[os.PathLike[str] | str] | os.PathLike[str] | str,
    output_dir: os.PathLike[str] | str,
    keywords: Mapping[str, Sequence[str] | str] | Sequence[str] | None = None,
) -> dict[str, Any]:
    """Inspect local product files and write deterministic audit artifacts.

    The returned mapping contains the same three document objects that are
    written to ``manifest.json``, ``extraction.json`` and ``review_queue.json``.
    No timestamp, random identifier, network result, or unscoped OCR output is
    included, so identical inputs at identical paths yield byte-stable files.
    """

    destination = Path(os.fspath(output_dir)).expanduser().resolve()
    input_files = _collect_input_files(paths, destination)
    section_keywords = _keyword_map(keywords)

    manifest_files: list[dict[str, Any]] = []
    extracted_records: list[dict[str, Any]] = []
    review_items: list[dict[str, Any]] = []

    for path in input_files:
        fingerprint = fingerprint_file(path)
        raw_records, issues = _extract_path(path)
        file_records: list[dict[str, Any]] = []
        file_review_items: list[dict[str, Any]] = []
        preflight_findings = [
            finding
            for raw_record in raw_records
            for finding in detect_prohibited_customer_data(
                str(raw_record.get("raw_text", ""))
            )
        ]
        coordinate_index = {
            (
                record.get("location", {}).get("sheet"),
                record.get("location", {}).get("row"),
                record.get("location", {}).get("column"),
            ): str(record.get("raw_text", "")).strip()
            for record in raw_records
            if record.get("location", {}).get("row") is not None
            and record.get("location", {}).get("column") is not None
        }
        for raw_record in raw_records:
            label = unicodedata.normalize(
                "NFKC", str(raw_record.get("raw_text", ""))
            ).strip()
            if not re.fullmatch(r"被保险人\s*年龄|insured\s*age", label, re.IGNORECASE):
                continue
            location = raw_record.get("location", {})
            neighbor = coordinate_index.get(
                (
                    location.get("sheet"),
                    location.get("row"),
                    int(location.get("column", 0)) + 1,
                )
            )
            if neighbor and re.fullmatch(
                r"\d{1,3}(?:\s*(?:周?岁|years?))?", neighbor, re.IGNORECASE
            ):
                preflight_findings.append(
                    {
                        "end": len(label),
                        "match_sha256": _sha256_text(label),
                        "start": 0,
                        "type": "customer_field",
                    }
                )
        if preflight_findings:
            file_review_items.append(
                _review_item(
                    source_sha256=str(fingerprint["sha256"]),
                    route="reject_prohibited_customer_data",
                    reason="The entire source was quarantined before record output because it contains customer-specific data.",
                    detection_types=[finding["type"] for finding in preflight_findings],
                )
            )
        else:
            for raw_record in raw_records:
                record, record_reviews = _enrich_record(
                    raw_record, fingerprint, section_keywords
                )
                file_records.append(record)
                file_review_items.extend(record_reviews)
            file_review_items.extend(
                _issue_review_item(issue, str(fingerprint["sha256"]))
                for issue in issues
            )

        extractors = sorted(
            {
                (_canonical_json(record["extractor"]), record["extractor"]["name"])
                for record in file_records
            }
        )
        manifest_entry = dict(fingerprint)
        if preflight_findings:
            manifest_entry["name"] = "[REDACTED:REJECTED_SOURCE]"
            manifest_entry["path"] = "[REDACTED:REJECTED_SOURCE]"
        manifest_entry.update(
            {
                "extractors": [json.loads(serialized) for serialized, _ in extractors],
                "inspection_status": "rejected_customer_data"
                if preflight_findings
                else "review_required"
                if file_review_items
                else "accepted",
                "prohibited_customer_data_records": len(preflight_findings),
                "record_count": len(file_records),
                "review_item_count": len(file_review_items),
            }
        )
        manifest_files.append(manifest_entry)
        extracted_records.extend(file_records)
        review_items.extend(file_review_items)

    manifest_files.sort(key=lambda item: (item["path"], item["sha256"]))
    extracted_records.sort(
        key=lambda item: (
            item["source_path"],
            _canonical_json(item["location"]),
            item["record_type"],
            item["raw_text_sha256"],
            item["id"],
        )
    )
    review_items.sort(
        key=lambda item: (
            item["source_sha256"],
            _canonical_json(item["location"]),
            item["route"],
            item["id"],
        )
    )

    manifest = {
        "files": manifest_files,
        "schema_version": SCHEMA_VERSION,
        "supported_extensions": sorted(_SUPPORTED_SUFFIXES),
    }
    extraction = {
        "records": extracted_records,
        "schema_version": SCHEMA_VERSION,
    }
    review_queue = {
        "items": review_items,
        "llm_called": False,
        "schema_version": SCHEMA_VERSION,
    }

    manifest_path = destination / "manifest.json"
    extraction_path = destination / "extraction.json"
    review_queue_path = destination / "review_queue.json"
    _write_stable_json(manifest_path, manifest)
    _write_stable_json(extraction_path, extraction)
    _write_stable_json(review_queue_path, review_queue)

    return {
        "extraction": extraction,
        "manifest": manifest,
        "output_files": {
            "extraction": str(extraction_path),
            "manifest": str(manifest_path),
            "review_queue": str(review_queue_path),
        },
        "review_queue": review_queue,
    }


__all__ = [
    "detect_prohibited_customer_data",
    "fingerprint_file",
    "inspect_inputs",
    "route_confidence",
]
